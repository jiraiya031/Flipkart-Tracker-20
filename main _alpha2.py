# Importing libraries
import requests
from datetime import datetime
from bs4 import BeautifulSoup
import smtplib
from email.message import EmailMessage
import time
from keep_alive import keep_alive
import pytz
from openpyxl import load_workbook
from openpyxl import Workbook



keep_alive()
filename = 'Price_Tracker.xlsx'
price_black =[]
price_blue =[]
time_stamp = []
try:
    # Load the existing workbook
    workbook = load_workbook(filename)
    # Select the active worksheet
    worksheet = workbook.active
    # Determine the row to start appending new data
    start_row = worksheet.max_row + 1
except FileNotFoundError:
    # Create a new workbook and worksheet if the file doesn't exist
    workbook = Workbook()
    worksheet = workbook.active
    print("book created")
    # Add headers to the worksheet
    worksheet.cell(row=1, column=1, value='Product1 Price')
    worksheet.cell(row=1, column=2, value='Product2 Price')
    worksheet.cell(row=1, column=3, value='Time')
    # Start appending new data from the second row
    start_row = 2


# Adding column headers to the worksheet

    # Start appending new data from the second row
   


# Setting up a list of products to track
products = [
    {
        "name": "SONY WH-XB910N Black",
        "url": "https://www.flipkart.com/sony-wh-xb910n-active-noise-cancellation-enabled-bluetooth-headset/p/itmd6a2096146f4d?pid=ACCGBHRAFRJAPTSX&lid",
        "target_price": 12000
    },
    {
        "name": "SONY WH-XB910N Blue",
        "url": "https://www.flipkart.com/sony-wh-xb910n-active-noise-cancellation-enabled-bluetooth-headset/p/itm668d0bbb72e6b",
        "target_price": 12000
    }
    
]

# Setting up email details
sender_email = "satyam1005@yahoo.in"
sender_password = "xpgpnszxnsbwfslw"
receiver_email = "gsatyam1005@gmail.com"

# Creating an email message object
msg = EmailMessage()
msg["Subject"] = "Price Drop Alert"
msg["From"] = sender_email
msg["To"] = receiver_email

# Creating a function to send email
def send_email(product):
    # Setting up the email content
    msg.set_content(f"The price of {product['name']} has dropped below {product['target_price']}. Here is the link: {product['url']}")
    # Sending the email using SMTP and TLS
    with smtplib.SMTP("smtp.mail.yahoo.com", 587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)
    # Printing a confirmation message
    print(f"Email sent for {product['name']}")

# Creating a function to check price
def check_price(product,n):
    count = n
    # Getting the product page using requests
    response = requests.get(product["url"])
    # Parsing the HTML using BeautifulSoup
    soup = BeautifulSoup(response.text, "html.parser")
    #print(soup.prettify())
    # Finding the price element using class name
    price_element = soup.find("div", {"class": "_30jeq3 _16Jk6d"}).get_text(strip=True)

    # Extracting the price text and removing commas and rupee symbol
    price_element = price_element.replace(",", "").replace("₹", "")
    # Converting the price text to a float value
    price_value = float(price_element.replace(",", ""))
    # Comparing the price value with the target price
    if price_value < product["target_price"]:
        # Sending an email if the price is lower than the target
        send_email(product)
        # Returning True to indicate that the email was sent
        return True
    else:
        # Printing a message if the price is higher than the target
        print(f"The price of {product['name']} is still {price_value}")
        
       
        tz_IN = pytz.timezone('Asia/Calcutta') 

        # Get the current time in India
        datetime_IN = datetime.now(tz_IN)
        datetime_IN = datetime_IN.strftime('%Y-%m-%d %H:%M:%S')
        # Format the time as a string and print it
        print("Time:", datetime_IN)
        if (count%2) == 0:
          price_blue.append(price_value)
 
        else :
         price_black.append(price_value) 
         time_stamp.append(datetime_IN)

        print(price_black)
        print(price_blue)
        print(time_stamp)
    for i, (price1, price2, time_data) in enumerate(zip(price_black, price_blue, time_stamp)):
        worksheet.cell(row=start_row+i, column=1, value=price1)
        worksheet.cell(row=start_row+i, column=2, value=price2)
        worksheet.cell(row=start_row+i, column=3, value=time_data)
    workbook.save(filename)
        # Returning False to indicate that the email was not sent
    return False



n=0
# Creating a while loop to run the code repeatedly
while True:
    # Looping through each product in the list
    for product in products:
        n=n+1
        # Checking the price of the product and storing the result
        result = check_price(product,n)
        # Removing the product from the list if the email was sent
        if result:
            products.remove(product)
    
    time.sleep(1000)
    # Breaking the loop if there are no more products to check
    if len(products) == 0:
        print(len(products))
        print("no products")
        break